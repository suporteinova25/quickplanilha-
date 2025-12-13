import os
import io
import re
from flask import Flask, render_template_string, request, redirect, session, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
import pg8000
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)

# ==================== CONFIGURAÇÕES ====================
DB_HOST = "mdm.inovaptt.com.br"
DB_PORT = 5432
DB_NAME = "hmdm"
# =======================================================

def normalize(m):
    if not m:
        return ""
    m = m.strip().upper()
    if m.endswith("_O"):
        m = m[:-2]
    if m.startswith("TELOX_") or m.startswith("TELO_"):
        if "_" in m:
            m = m.split("_", 1)[1]
    return m

def col_c(model_n):
    mu = model_n.upper()
    if mu.startswith("TE") or mu == "RG750" or mu.startswith("IS530"):
        return "HT"
    if mu.startswith("SM-") or mu == "RG935":
        return "TABLET"
    return ""

@app.route("/", methods=["GET", "POST"])
def login():
    form_html = """
    <h2>Login HMDM</h2>
    <form method="post">
        Usuário:<br><input name="user" required autocomplete="off"><br><br>
        Senha:<br><input type="password" name="pwd" required><br><br>
        <button>Entrar</button>
    </form>
    {% if msg %}<br><small style="color:red">{{msg}}</small>{% endif %}
    """
    if request.method == "POST":
        user = request.form["user"]
        pwd = request.form["pwd"]
        try:
            with pg8000.connect(host=DB_HOST, port=DB_PORT, database=DB_NAME,
                                user=user, password=pwd) as conn:
                pass
            session["authenticated"] = True
            session["user"] = user
            session["pwd"] = pwd
            return redirect("/busca")
        except Exception as e:
            return render_template_string(form_html, msg="Erro de login: " + str(e))
    return render_template_string(form_html)

@app.route("/busca", methods=["GET", "POST"])
def busca():
    if not session.get("authenticated"):
        return redirect("/")

    if request.method == "POST":
        sufixo = request.form["sufixo"].strip()
        if not sufixo:
            return "Por favor, informe pelo menos um caractere para busca."

        try:
            with pg8000.connect(host=DB_HOST, port=DB_PORT, database=DB_NAME,
                                user=session["user"], password=session["pwd"]) as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT
                            "number",
                            NULLIF(TRIM(info::json->>'imei'), '') AS imei1,
                            NULLIF(TRIM(info::json->>'iccid'), '') AS iccid1,
                            NULLIF(TRIM(info::json->>'phone'), '') AS linha1,
                            COALESCE(
                                NULLIF(TRIM(info::json->>'imei2'), ''),
                                NULLIF(TRIM(info::json->>'imeiSlot2'), '')
                            ) AS imei2,
                            COALESCE(
                                NULLIF(TRIM(info::json->>'iccid2'), ''),
                                NULLIF(TRIM(info::json->>'iccidSlot2'), '')
                            ) AS iccid2,
                            COALESCE(
                                NULLIF(TRIM(info::json->>'phone2'), ''),
                                NULLIF(TRIM(info::json->>'line2'), ''),
                                NULLIF(TRIM(info::json->>'number2'), '')
                            ) AS linha2,
                            (info::json->>'model') AS model,
                            (info::json->>'serial') AS serial
                        FROM devices
                        WHERE "number" ILIKE %s
                           OR info::json->>'phone2' ILIKE %s
                           OR info::json->>'line2' ILIKE %s
                        ORDER BY "number"
                    """, (f"%{sufixo}%", f"%{sufixo}%", f"%{sufixo}%"))
                    rows = cur.fetchall()

            if not rows:
                return "Nenhum dispositivo encontrado com esse sufixo."

            wb = load_workbook("modelo.xlsx")
            ws = wb.active

            # Limpa colunas B até M a partir da linha 2
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=13):
                for cell in row:
                    cell.value = None
                    cell.font = Font()

            roxo = Font(color="9C27B0")
            vermelho = Font(color="F44336")
            azul = Font(color="1976D2")
            vermelho_fundo = PatternFill(start_color="FF0000", fill_type="solid")

            for idx, row in enumerate(rows, start=2):
                number, imei1, iccid1, linha1, imei2, iccid2, linha2, model_raw, serial = row
                model = normalize(model_raw)

                ws[f"B{idx}"] = model
                ws[f"C{idx}"] = col_c(model)
                ws[f"D{idx}"] = number or ""
                ws[f"G{idx}"] = imei1 or ""
                ws[f"H{idx}"] = iccid1 or ""

                if iccid1:
                    iccid_str = str(iccid1)
                    if iccid_str.startswith("895510"):
                        ws[f"H{idx}"].font = roxo
                    elif iccid_str.startswith("8955053"):
                        ws[f"H{idx}"].font = vermelho
                    else:
                        ws[f"H{idx}"].font = azul

                # Correções anteriores mantidas
                ws[f"I{idx}"] = linha1 if linha1 else ""   # Só se tiver linha real
                ws[f"J{idx}"] = imei2 if imei2 else ""     # Só se tiver segundo IMEI

                ws[f"K{idx}"] = iccid2 or ""
                if iccid2:
                    iccid_str = str(iccid2)
                    if iccid_str.startswith("895510"):
                        ws[f"K{idx}"].font = roxo
                    elif iccid_str.startswith("8955053"):
                        ws[f"K{idx}"].font = vermelho
                    else:
                        ws[f"K{idx}"].font = azul

                ws[f"L{idx}"] = linha2 or ""
                ws[f"M{idx}"] = serial or ""

            ultima_linha = len(rows) + 1

            # IMEI1 (G): destaca duplicados (ignora vazios com AND)
            ws.conditional_formatting.add(f"G2:G{ultima_linha}",
                FormulaRule(formula=[f'AND(G2<>"", COUNTIF($G$2:$G${ultima_linha},G2)>1)'],
                            fill=vermelho_fundo))

            # IMEI2 (J): destaca apenas duplicados reais (não pinta células vazias)
            ws.conditional_formatting.add(f"J2:J{ultima_linha}",
                FormulaRule(formula=[f'COUNTIF($J$2:$J${ultima_linha},J2)>1'],
                            fill=vermelho_fundo))

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            safe_sufixo = re.sub(r'[^\w\-]', '_', sufixo)

            return send_file(
                buf,
                as_attachment=True,
                download_name=f"dispositivos_{safe_sufixo}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            return f"Erro ao conectar ou processar dados: {str(e)}<br><a href='/logout'>Sair</a>"

    return """
    <h2>Busca Dispositivos Dual-Chip</h2>
    <form method="post">
        Sufixo da linha: <input name="sufixo" size="35" required placeholder="ex: 99, 123 ou 11999999999">
        <button style="padding:10px">Gerar Planilha Excel</button>
    </form>
    <br><small>Agora aceita qualquer quantidade de dígitos (mínimo 1)</small><br><br>
    <a href="/logout">Sair</a>
    """

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)